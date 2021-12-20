# -->>>清安<<<---
from openpyxl import load_workbook
from uconfig.readconfig import readConfig

def load_excel(file):
    """
    用于读取测试用例所在的表格
    """
    _data = load_workbook(file)
    _value = _data.active
    return _value


def parse_case(file):
    """
    用于将excel中的测试用例转为pytest可识别的测试用例
    """
    excel_value = load_excel(file)
    _cases = []
    for i in range(2, excel_value.max_row + 1):
        _cases.append({'ele': excel_value['B' + str(i)].value, 'location_method': excel_value['C' + str(i)].value,
                       'action': excel_value['D' + str(i)].value, 'value': excel_value['E' + str(i)].value,
                       'loc': excel_value['F' + str(i)].value})

    return _cases
